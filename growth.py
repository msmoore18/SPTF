import streamlit as st
import pandas as pd
import plotly.express as px
import msoffcrypto
import io
import openpyxl

# Streamlit page configuration
st.set_page_config(layout="wide")

# Load password-protected Excel file
def load_data(password):
    file_path = "SPTF_Count.xlsx"
    try:
        with open(file_path, "rb") as file:
            decrypted = io.BytesIO()
            office_file = msoffcrypto.OfficeFile(file)

            if not office_file.is_encrypted():
                return pd.ExcelFile(file_path)

            office_file.load_key(password)
            office_file.decrypt(decrypted)
            return pd.ExcelFile(decrypted)
    except Exception:
        return None

# Ensure session state variables exist
if "password_input" not in st.session_state:
    st.session_state["password_input"] = ""

if "data" not in st.session_state:
    st.title("Enter Password to Access Data")

    def submit_password():
        password = st.session_state.password_input
        if password:
            data = load_data(password)
            if data is None:
                st.session_state["password_error"] = True
            else:
                st.session_state["data"] = data
                st.session_state["password"] = password  # Store password in session
                st.session_state["password_error"] = False

    # Initialize the error state
    if "password_error" not in st.session_state:
        st.session_state["password_error"] = False

    st.text_input(
        "Excel File Password:",
        type="password",
        key="password_input",
        on_change=submit_password
    )

    if st.button("Enter"):
        submit_password()

    # Handle errors or success after callback/button press
    if st.session_state.get("password_error"):
        st.error("Incorrect password or file issue. Please try again.")
    elif "data" in st.session_state:
        st.rerun()

    st.stop()

# Ensure data is available before proceeding
if "data" in st.session_state and "password" in st.session_state:
    xls = st.session_state["data"]
    excel_password = st.session_state["password"]  # Retrieve stored password
else:
    st.stop()

# Sidebar navigation
st.sidebar.title("Navigation")
page = st.sidebar.radio("Go to", ["Lot Map", "Tree Inventory", "Projected Tree Inventory", "Tree Maintenance"])

if page == "Projected Tree Inventory":
    st.title("Projected Tree Inventory")

    # Load the "calculations" sheet with decryption
    file_path = "SPTF_Count.xlsx"
    
    with open(file_path, "rb") as file:
        decrypted = io.BytesIO()
        office_file = msoffcrypto.OfficeFile(file)
        office_file.load_key(excel_password)  # Use stored password
        office_file.decrypt(decrypted)
    
    wb = openpyxl.load_workbook(decrypted, data_only=True)
    ws = wb["calculations"]

    # Retrieve Tree Heights from Column N (N2:N28)
    tree_heights = {f"O{row}": ws[f"N{row}"].value for row in range(2, 29)}

    # User input for B30
    user_input_b30 = st.number_input("How many new trees do you want to buy each year?", value=ws["B30"].value or 0)

    st.write("### Modify Projected Inventory Using Sliders")

    if "cell_modifications" not in st.session_state:
        st.session_state["cell_modifications"] = {}  # Stores modified values

    # Create a grid layout for the sliders
    columns = st.columns(3)  # Display sliders in 3 columns for a clean UI

    for i, (cell, height) in enumerate(tree_heights.items()):
        slider_value = st.session_state["cell_modifications"].get(cell, ws[cell].value or 0)

        # Distribute sliders across columns
        with columns[i % 3]:  
            new_value = st.slider(
                label=f"{height} ft Trees",
                min_value=0,
                max_value=500,  # Limit max input to 500
                value=int(slider_value),
                key=cell
            )
        
        # Store modified values in session state
        st.session_state["cell_modifications"][cell] = new_value

    # Button to apply all changes to Excel
    if st.button("Update Calculations"):
        ws["B30"].value = user_input_b30

        # Apply all saved changes to selected tree heights
        for cell, value in st.session_state["cell_modifications"].items():
            ws[cell].value = value

        # Save the updated file
        excel_modified = io.BytesIO()
        wb.save(excel_modified)
        excel_modified.seek(0)

        # Reload updated calculations
        updated_wb = openpyxl.load_workbook(excel_modified, data_only=True)
        updated_ws = updated_wb["calculations"]

        # Extract the A1:L28 range as a DataFrame
        data_range = []
        for row in updated_ws.iter_rows(min_row=1, max_row=28, min_col=1, max_col=12, values_only=True):
            data_range.append(row)
        
        df_calculations = pd.DataFrame(data_range)

        st.success("Excel file updated successfully!")

        # Display updated table without first row
        st.write("### Updated Calculations Table")
        st.dataframe(df_calculations.iloc[1:], hide_index=True)

        # Provide a download option for the updated Excel file
        st.download_button(
            label="Download Updated Excel",
            data=excel_modified,
            file_name="Updated_SPTF_Count.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # 📊 Bar Chart: Display user modifications as Tree Count vs Tree Height
    all_tree_counts = {
        height: st.session_state["cell_modifications"].get(cell, ws[cell].value) or 0
        for cell, height in tree_heights.items()
    }

    df_chart = pd.DataFrame({
        "Tree Height (ft)": list(all_tree_counts.keys()),
        "Tree Count": list(all_tree_counts.values())
    })

    if not df_chart.empty:
        fig = px.bar(df_chart, x="Tree Height (ft)", y="Tree Count", title="Projected Tree Inventory",
                     labels={"Tree Count": "Number of Trees", "Tree Height (ft)": "Tree Height (ft)"})
        st.plotly_chart(fig)
